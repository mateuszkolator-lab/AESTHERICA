from fastapi import APIRouter, Depends, Header
from typing import Optional
from datetime import datetime, timezone
from io import BytesIO
import os
import openpyxl
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse

from models.database import get_db
from routers.auth import verify_token

router = APIRouter(tags=["Stats & Dashboard"])

def get_auth(authorization: str = Header(None)):
    return verify_token(authorization)

@router.get("/stats")
async def get_stats(year: Optional[int] = None, location_id: Optional[str] = None, user: dict = Depends(get_auth)):
    db = get_db()
    
    if not year:
        year = datetime.now(timezone.utc).year
    
    year_str = str(year)
    
    # Base query filter
    base_filter = {}
    if location_id:
        base_filter["location_id"] = location_id
    
    # Total patients
    total = await db.patients.count_documents(base_filter)
    
    # By status
    by_status = {}
    for status in ["consultation", "planned", "awaiting", "operated"]:
        by_status[status] = await db.patients.count_documents({**base_filter, "status": status})
    
    # Procedures by month (operated patients with surgery_date in given year)
    proc_match = {**base_filter, "surgery_date": {"$regex": f"^{year_str}"}, "status": "operated"}
    procedures_pipeline = [
        {"$match": proc_match},
        {"$group": {
            "_id": {"$substr": ["$surgery_date", 5, 2]},
            "count": {"$sum": 1}
        }},
        {"$project": {"_id": 0, "month": "$_id", "count": 1}}
    ]
    procedures_by_month = await db.patients.aggregate(procedures_pipeline).to_list(12)
    
    # Revenue by month - OPERATED
    rev_operated_match = {**base_filter, "surgery_date": {"$regex": f"^{year_str}"}, "status": "operated", "price": {"$gt": 0}}
    revenue_operated_pipeline = [
        {"$match": rev_operated_match},
        {"$group": {
            "_id": {"$substr": ["$surgery_date", 5, 2]},
            "revenue": {"$sum": "$price"},
            "count": {"$sum": 1}
        }},
        {"$project": {"_id": 0, "month": "$_id", "revenue": 1, "count": 1}}
    ]
    revenue_by_month_operated = await db.patients.aggregate(revenue_operated_pipeline).to_list(12)
    
    # Revenue by month - PLANNED
    rev_planned_match = {**base_filter, "surgery_date": {"$regex": f"^{year_str}"}, "status": "planned", "price": {"$gt": 0}}
    revenue_planned_pipeline = [
        {"$match": rev_planned_match},
        {"$group": {
            "_id": {"$substr": ["$surgery_date", 5, 2]},
            "revenue": {"$sum": "$price"},
            "count": {"$sum": 1}
        }},
        {"$project": {"_id": 0, "month": "$_id", "revenue": 1, "count": 1}}
    ]
    revenue_by_month_planned = await db.patients.aggregate(revenue_planned_pipeline).to_list(12)
    
    # Combined revenue (for backward compat)
    revenue_by_month = revenue_by_month_operated
    
    # Procedures by location
    locations = await db.locations.find({}, {"_id": 0}).to_list(100)
    location_map = {loc["id"]: loc["name"] for loc in locations}
    
    location_pipeline = [
        {"$match": {**base_filter, "status": "operated", "location_id": {"$ne": None}}},
        {"$group": {"_id": "$location_id", "count": {"$sum": 1}}},
        {"$project": {"_id": 0, "location_id": "$_id", "count": 1}}
    ]
    location_data = await db.patients.aggregate(location_pipeline).to_list(100)
    procedures_by_location = [
        {"location": location_map.get(item["location_id"], "Nieznana"), "count": item["count"]}
        for item in location_data
    ]
    
    # Revenue by location
    rev_loc_pipeline = [
        {"$match": {"surgery_date": {"$regex": f"^{year_str}"}, "status": {"$in": ["planned", "operated"]}, "location_id": {"$ne": None}, "price": {"$gt": 0}}},
        {"$group": {
            "_id": {"location_id": "$location_id", "status": "$status"},
            "revenue": {"$sum": "$price"},
            "count": {"$sum": 1}
        }},
        {"$project": {"_id": 0, "location_id": "$_id.location_id", "status": "$_id.status", "revenue": 1, "count": 1}}
    ]
    rev_loc_data = await db.patients.aggregate(rev_loc_pipeline).to_list(100)
    revenue_by_location = [
        {**item, "location": location_map.get(item["location_id"], "Nieznana")}
        for item in rev_loc_data
    ]
    
    return {
        "total_patients": total,
        "by_status": by_status,
        "procedures_by_month": procedures_by_month,
        "revenue_by_month": revenue_by_month,
        "revenue_by_month_operated": revenue_by_month_operated,
        "revenue_by_month_planned": revenue_by_month_planned,
        "procedures_by_location": procedures_by_location,
        "revenue_by_location": revenue_by_location,
        "locations": [{"id": loc["id"], "name": loc["name"]} for loc in locations]
    }

@router.get("/dashboard")
async def get_dashboard(user: dict = Depends(get_auth)):
    db = get_db()
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # All upcoming surgeries (from today onwards)
    upcoming = await db.patients.find(
        {"surgery_date": {"$gte": today}, "status": {"$in": ["planned", "awaiting"]}},
        {"_id": 0}
    ).sort("surgery_date", 1).to_list(100)
    
    # Recent patients
    recent = await db.patients.find({}, {"_id": 0}).sort("created_at", -1).to_list(5)
    
    # Quick stats
    total = await db.patients.count_documents({})
    operated = await db.patients.count_documents({"status": "operated"})
    planned = await db.patients.count_documents({"status": "planned"})
    awaiting = await db.patients.count_documents({"status": "awaiting"})
    consultation = await db.patients.count_documents({"status": "consultation"})
    
    return {
        "upcoming_surgeries": upcoming,
        "recent_patients": recent,
        "stats": {
            "total": total,
            "operated": operated,
            "planned": planned,
            "awaiting": awaiting,
            "consultation": consultation
        }
    }

@router.get("/export/patients")
async def export_patients(
    status: Optional[str] = None,
    year: Optional[int] = None,
    user: dict = Depends(get_auth)
):
    db = get_db()
    query = {}
    if status:
        query["status"] = status
    if year:
        query["surgery_date"] = {"$regex": f"^{year}"}
    
    patients = await db.patients.find(query, {"_id": 0}).to_list(10000)
    locations = await db.locations.find({}, {"_id": 0}).to_list(100)
    location_map = {loc["id"]: loc["name"] for loc in locations}
    
    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"
    
    # Headers
    headers = ["First Name", "Last Name", "Email", "Phone", "Status", "Procedure", "Surgery Date", "Location", "Price", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
    
    # Data
    for row, patient in enumerate(patients, 2):
        ws.cell(row=row, column=1, value=patient.get("first_name", ""))
        ws.cell(row=row, column=2, value=patient.get("last_name", ""))
        ws.cell(row=row, column=3, value=patient.get("email", ""))
        ws.cell(row=row, column=4, value=patient.get("phone", ""))
        ws.cell(row=row, column=5, value=patient.get("status", ""))
        ws.cell(row=row, column=6, value=patient.get("procedure_type", ""))
        ws.cell(row=row, column=7, value=patient.get("surgery_date", ""))
        ws.cell(row=row, column=8, value=location_map.get(patient.get("location_id"), ""))
        ws.cell(row=row, column=9, value=patient.get("price", ""))
        ws.cell(row=row, column=10, value=patient.get("notes", ""))
    
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"patients_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/stats/waiting-summary")
async def get_waiting_summary(user: dict = Depends(get_auth)):
    """Get summary of patients waiting by status and procedure type"""
    db = get_db()
    
    # Aggregate patients by status and procedure_type (excluding 'operated')
    pipeline = [
        {"$match": {"status": {"$in": ["consultation", "planned", "awaiting"]}}},
        {"$group": {
            "_id": {
                "status": "$status",
                "procedure_type": {"$ifNull": ["$procedure_type", "Nieokreślony"]}
            },
            "count": {"$sum": 1}
        }},
        {"$project": {
            "_id": 0,
            "status": "$_id.status",
            "procedure_type": "$_id.procedure_type",
            "count": 1
        }},
        {"$sort": {"count": -1}}
    ]
    
    results = await db.patients.aggregate(pipeline).to_list(1000)
    return results
