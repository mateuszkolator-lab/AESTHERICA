from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import base64
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Settings
JWT_SECRET = os.environ.get('JWT_SECRET', 'aestheticamd-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# App password (simple single-user auth)
APP_PASSWORD = os.environ.get('APP_PASSWORD', 'doctor2024')

# Create the main app
app = FastAPI(title="AestheticaMD API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class LoginRequest(BaseModel):
    password: str

class TokenResponse(BaseModel):
    token: str
    expires_at: str

class Location(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    address: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class LocationCreate(BaseModel):
    name: str
    address: Optional[str] = None

class ProcedureType(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    default_price: Optional[float] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ProcedureTypeCreate(BaseModel):
    name: str
    description: Optional[str] = None
    default_price: Optional[float] = None

class SurgerySlot(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    location_id: Optional[str] = None
    max_patients: int = 1
    notes: Optional[str] = None
    assigned_patient_id: Optional[str] = None
    is_full: bool = False
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class SurgerySlotCreate(BaseModel):
    date: str
    location_id: Optional[str] = None
    max_patients: int = 1
    notes: Optional[str] = None
    is_full: bool = False

class Photo(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    data: str  # Base64 encoded
    filename: str
    caption: Optional[str] = None
    category: Optional[str] = None  # before, after, during, other
    uploaded_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Visit(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    type: str  # consultation, surgery, follow_up
    notes: Optional[str] = None
    photos: List[Photo] = []
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Patient(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    first_name: str
    last_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    date_of_birth: Optional[str] = None
    gender: Optional[str] = None
    status: str = "consultation"  # consultation, planned, awaiting, operated
    procedure_type: Optional[str] = None
    preferred_date_start: Optional[str] = None
    preferred_date_end: Optional[str] = None
    surgery_date: Optional[str] = None
    location_id: Optional[str] = None
    price: Optional[float] = None
    notes: Optional[str] = None
    visits: List[Visit] = []
    google_calendar_event_id: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class PatientCreate(BaseModel):
    first_name: str
    last_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    date_of_birth: Optional[str] = None
    gender: Optional[str] = None
    status: str = "consultation"
    procedure_type: Optional[str] = None
    preferred_date_start: Optional[str] = None
    preferred_date_end: Optional[str] = None
    surgery_date: Optional[str] = None
    location_id: Optional[str] = None
    price: Optional[float] = None
    notes: Optional[str] = None

class PatientUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    date_of_birth: Optional[str] = None
    gender: Optional[str] = None
    status: Optional[str] = None
    procedure_type: Optional[str] = None
    preferred_date_start: Optional[str] = None
    preferred_date_end: Optional[str] = None
    surgery_date: Optional[str] = None
    location_id: Optional[str] = None
    price: Optional[float] = None
    notes: Optional[str] = None

class VisitCreate(BaseModel):
    date: str
    type: str
    notes: Optional[str] = None

class PhotoUpload(BaseModel):
    data: str
    filename: str
    caption: Optional[str] = None
    category: Optional[str] = None

class StatsResponse(BaseModel):
    total_patients: int
    by_status: dict
    procedures_by_month: List[dict]
    procedures_by_location: List[dict]
    revenue_by_month: List[dict]

# ==================== AUTH HELPERS ====================

def create_token(data: dict) -> str:
    expires = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    to_encode = data.copy()
    to_encode.update({"exp": expires})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(request: LoginRequest):
    if request.password != APP_PASSWORD:
        raise HTTPException(status_code=401, detail="Invalid password")
    
    token = create_token({"user": "doctor", "role": "admin"})
    expires = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    return TokenResponse(token=token, expires_at=expires.isoformat())

@api_router.get("/auth/verify")
async def verify_auth(user: dict = Depends(verify_token)):
    return {"valid": True, "user": user}

# ==================== LOCATION ROUTES ====================

@api_router.get("/locations", response_model=List[Location])
async def get_locations(user: dict = Depends(verify_token)):
    locations = await db.locations.find({}, {"_id": 0}).to_list(100)
    return locations

@api_router.post("/locations", response_model=Location)
async def create_location(location: LocationCreate, user: dict = Depends(verify_token)):
    loc = Location(**location.model_dump())
    await db.locations.insert_one(loc.model_dump())
    return loc

@api_router.put("/locations/{location_id}", response_model=Location)
async def update_location(location_id: str, location: LocationCreate, user: dict = Depends(verify_token)):
    result = await db.locations.find_one_and_update(
        {"id": location_id},
        {"$set": location.model_dump()},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Location not found")
    del result["_id"]
    return result

@api_router.delete("/locations/{location_id}")
async def delete_location(location_id: str, user: dict = Depends(verify_token)):
    result = await db.locations.delete_one({"id": location_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Location not found")
    return {"message": "Location deleted"}

# ==================== PROCEDURE TYPE ROUTES ====================

@api_router.get("/procedure-types", response_model=List[ProcedureType])
async def get_procedure_types(user: dict = Depends(verify_token)):
    procedure_types = await db.procedure_types.find({}, {"_id": 0}).to_list(100)
    return procedure_types

@api_router.post("/procedure-types", response_model=ProcedureType)
async def create_procedure_type(procedure_type: ProcedureTypeCreate, user: dict = Depends(verify_token)):
    pt = ProcedureType(**procedure_type.model_dump())
    await db.procedure_types.insert_one(pt.model_dump())
    return pt

@api_router.put("/procedure-types/{procedure_type_id}", response_model=ProcedureType)
async def update_procedure_type(procedure_type_id: str, procedure_type: ProcedureTypeCreate, user: dict = Depends(verify_token)):
    result = await db.procedure_types.find_one_and_update(
        {"id": procedure_type_id},
        {"$set": procedure_type.model_dump()},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Procedure type not found")
    del result["_id"]
    return result

@api_router.delete("/procedure-types/{procedure_type_id}")
async def delete_procedure_type(procedure_type_id: str, user: dict = Depends(verify_token)):
    result = await db.procedure_types.delete_one({"id": procedure_type_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Procedure type not found")
    return {"message": "Procedure type deleted"}

# ==================== SURGERY SLOTS ROUTES ====================

@api_router.get("/surgery-slots", response_model=List[SurgerySlot])
async def get_surgery_slots(
    include_past: bool = False,
    user: dict = Depends(verify_token)
):
    query = {}
    if not include_past:
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        query["date"] = {"$gte": today}
    
    slots = await db.surgery_slots.find(query, {"_id": 0}).sort("date", 1).to_list(200)
    return slots

@api_router.post("/surgery-slots", response_model=SurgerySlot)
async def create_surgery_slot(slot: SurgerySlotCreate, user: dict = Depends(verify_token)):
    slot_obj = SurgerySlot(**slot.model_dump())
    await db.surgery_slots.insert_one(slot_obj.model_dump())
    return slot_obj

@api_router.put("/surgery-slots/{slot_id}", response_model=SurgerySlot)
async def update_surgery_slot(slot_id: str, slot: SurgerySlotCreate, user: dict = Depends(verify_token)):
    result = await db.surgery_slots.find_one_and_update(
        {"id": slot_id},
        {"$set": slot.model_dump()},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Surgery slot not found")
    del result["_id"]
    return result

@api_router.delete("/surgery-slots/{slot_id}")
async def delete_surgery_slot(slot_id: str, user: dict = Depends(verify_token)):
    result = await db.surgery_slots.delete_one({"id": slot_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Surgery slot not found")
    return {"message": "Surgery slot deleted"}

@api_router.post("/surgery-slots/{slot_id}/toggle-full")
async def toggle_slot_full(slot_id: str, user: dict = Depends(verify_token)):
    """Toggle the full status of a surgery slot"""
    slot = await db.surgery_slots.find_one({"id": slot_id})
    if not slot:
        raise HTTPException(status_code=404, detail="Surgery slot not found")
    
    new_status = not slot.get("is_full", False)
    await db.surgery_slots.update_one(
        {"id": slot_id},
        {"$set": {"is_full": new_status}}
    )
    return {"message": "Slot status updated", "is_full": new_status}

@api_router.get("/surgery-slots/calendar-data")
async def get_calendar_data(user: dict = Depends(verify_token)):
    """Get all data needed for calendar drag & drop"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Get all future slots
    slots = await db.surgery_slots.find(
        {"date": {"$gte": today}},
        {"_id": 0}
    ).sort("date", 1).to_list(200)
    
    # Get unassigned patients (waiting for surgery)
    unassigned_patients = await db.patients.find(
        {
            "status": {"$in": ["consultation", "awaiting"]},
            "$or": [
                {"surgery_date": None},
                {"surgery_date": ""}
            ]
        },
        {"_id": 0}
    ).to_list(500)
    
    # Get locations
    locations = await db.locations.find({}, {"_id": 0}).to_list(100)
    location_map = {loc["id"]: loc["name"] for loc in locations}
    
    # Enrich slots with location names and assigned patient info
    enriched_slots = []
    for slot in slots:
        slot_data = {**slot, "location_name": location_map.get(slot.get("location_id"), "")}
        if slot.get("assigned_patient_id"):
            patient = await db.patients.find_one({"id": slot["assigned_patient_id"]}, {"_id": 0})
            if patient:
                slot_data["assigned_patient"] = {
                    "id": patient["id"],
                    "first_name": patient["first_name"],
                    "last_name": patient["last_name"],
                    "procedure_type": patient.get("procedure_type")
                }
        enriched_slots.append(slot_data)
    
    return {
        "slots": enriched_slots,
        "unassigned_patients": unassigned_patients,
        "locations": locations
    }

@api_router.get("/surgery-slots/suggestions")
async def get_slot_suggestions(user: dict = Depends(verify_token)):
    """Get surgery slots with suggested patients based on preferred dates"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Get all future unassigned slots
    slots = await db.surgery_slots.find(
        {"date": {"$gte": today}, "assigned_patient_id": None},
        {"_id": 0}
    ).sort("date", 1).to_list(100)
    
    # Get all patients waiting for surgery (consultation or awaiting status, no surgery date)
    waiting_patients = await db.patients.find(
        {
            "status": {"$in": ["consultation", "awaiting"]},
            "$or": [
                {"surgery_date": None},
                {"surgery_date": ""}
            ]
        },
        {"_id": 0}
    ).to_list(500)
    
    # Get locations for display
    locations = await db.locations.find({}, {"_id": 0}).to_list(100)
    location_map = {loc["id"]: loc["name"] for loc in locations}
    
    suggestions = []
    for slot in slots:
        slot_date = slot["date"]
        matching_patients = []
        
        for patient in waiting_patients:
            pref_start = patient.get("preferred_date_start")
            pref_end = patient.get("preferred_date_end")
            
            # Check if slot date falls within patient's preferred range
            if pref_start and pref_end:
                if pref_start <= slot_date <= pref_end:
                    matching_patients.append({
                        "id": patient["id"],
                        "first_name": patient["first_name"],
                        "last_name": patient["last_name"],
                        "procedure_type": patient.get("procedure_type"),
                        "preferred_date_start": pref_start,
                        "preferred_date_end": pref_end,
                        "phone": patient.get("phone"),
                        "match_score": 100  # Perfect match
                    })
            elif pref_start and not pref_end:
                if slot_date >= pref_start:
                    matching_patients.append({
                        "id": patient["id"],
                        "first_name": patient["first_name"],
                        "last_name": patient["last_name"],
                        "procedure_type": patient.get("procedure_type"),
                        "preferred_date_start": pref_start,
                        "preferred_date_end": pref_end,
                        "phone": patient.get("phone"),
                        "match_score": 80
                    })
            elif pref_end and not pref_start:
                if slot_date <= pref_end:
                    matching_patients.append({
                        "id": patient["id"],
                        "first_name": patient["first_name"],
                        "last_name": patient["last_name"],
                        "procedure_type": patient.get("procedure_type"),
                        "preferred_date_start": pref_start,
                        "preferred_date_end": pref_end,
                        "phone": patient.get("phone"),
                        "match_score": 80
                    })
        
        # Sort by match score
        matching_patients.sort(key=lambda x: x["match_score"], reverse=True)
        
        suggestions.append({
            "slot": {
                **slot,
                "location_name": location_map.get(slot.get("location_id"), "")
            },
            "suggested_patients": matching_patients[:10]  # Top 10 matches
        })
    
    return suggestions

@api_router.post("/surgery-slots/{slot_id}/assign/{patient_id}")
async def assign_patient_to_slot(slot_id: str, patient_id: str, user: dict = Depends(verify_token)):
    """Assign a patient to a surgery slot"""
    # Get the slot
    slot = await db.surgery_slots.find_one({"id": slot_id})
    if not slot:
        raise HTTPException(status_code=404, detail="Surgery slot not found")
    
    # Get the patient
    patient = await db.patients.find_one({"id": patient_id})
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    
    # Update the slot with assigned patient
    await db.surgery_slots.update_one(
        {"id": slot_id},
        {"$set": {"assigned_patient_id": patient_id}}
    )
    
    # Update patient with surgery date and status
    await db.patients.update_one(
        {"id": patient_id},
        {"$set": {
            "surgery_date": slot["date"],
            "status": "planned",
            "location_id": slot.get("location_id"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Patient assigned to surgery slot", "surgery_date": slot["date"]}

@api_router.post("/surgery-slots/{slot_id}/unassign")
async def unassign_patient_from_slot(slot_id: str, user: dict = Depends(verify_token)):
    """Remove patient assignment from a surgery slot"""
    slot = await db.surgery_slots.find_one({"id": slot_id})
    if not slot:
        raise HTTPException(status_code=404, detail="Surgery slot not found")
    
    patient_id = slot.get("assigned_patient_id")
    
    # Clear the slot assignment
    await db.surgery_slots.update_one(
        {"id": slot_id},
        {"$set": {"assigned_patient_id": None}}
    )
    
    # If there was a patient, clear their surgery date
    if patient_id:
        await db.patients.update_one(
            {"id": patient_id},
            {"$set": {
                "surgery_date": None,
                "status": "awaiting",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    return {"message": "Patient unassigned from surgery slot"}

# ==================== PATIENT ROUTES ====================

@api_router.get("/patients", response_model=List[Patient])
async def get_patients(
    status: Optional[str] = None,
    location_id: Optional[str] = None,
    preferred_date_start: Optional[str] = None,
    preferred_date_end: Optional[str] = None,
    surgery_date_start: Optional[str] = None,
    surgery_date_end: Optional[str] = None,
    sort_by: str = "created_at",
    sort_order: str = "desc",
    user: dict = Depends(verify_token)
):
    query = {}
    if status:
        query["status"] = status
    if location_id:
        query["location_id"] = location_id
    if preferred_date_start:
        query["preferred_date_start"] = {"$gte": preferred_date_start}
    if preferred_date_end:
        query["preferred_date_end"] = {"$lte": preferred_date_end}
    if surgery_date_start:
        query["surgery_date"] = {"$gte": surgery_date_start}
    if surgery_date_end:
        if "surgery_date" in query:
            query["surgery_date"]["$lte"] = surgery_date_end
        else:
            query["surgery_date"] = {"$lte": surgery_date_end}
    
    sort_direction = -1 if sort_order == "desc" else 1
    patients = await db.patients.find(query, {"_id": 0}).sort(sort_by, sort_direction).to_list(1000)
    return patients

@api_router.get("/patients/{patient_id}", response_model=Patient)
async def get_patient(patient_id: str, user: dict = Depends(verify_token)):
    patient = await db.patients.find_one({"id": patient_id}, {"_id": 0})
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    return patient

@api_router.post("/patients", response_model=Patient)
async def create_patient(patient: PatientCreate, user: dict = Depends(verify_token)):
    patient_obj = Patient(**patient.model_dump())
    await db.patients.insert_one(patient_obj.model_dump())
    return patient_obj

@api_router.put("/patients/{patient_id}", response_model=Patient)
async def update_patient(patient_id: str, patient: PatientUpdate, user: dict = Depends(verify_token)):
    update_data = {k: v for k, v in patient.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.patients.find_one_and_update(
        {"id": patient_id},
        {"$set": update_data},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Patient not found")
    del result["_id"]
    return result

@api_router.delete("/patients/{patient_id}")
async def delete_patient(patient_id: str, user: dict = Depends(verify_token)):
    result = await db.patients.delete_one({"id": patient_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Patient not found")
    return {"message": "Patient deleted"}

# ==================== VISIT ROUTES ====================

@api_router.post("/patients/{patient_id}/visits", response_model=Visit)
async def add_visit(patient_id: str, visit: VisitCreate, user: dict = Depends(verify_token)):
    visit_obj = Visit(**visit.model_dump())
    result = await db.patients.find_one_and_update(
        {"id": patient_id},
        {
            "$push": {"visits": visit_obj.model_dump()},
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        },
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Patient not found")
    return visit_obj

@api_router.put("/patients/{patient_id}/visits/{visit_id}")
async def update_visit(patient_id: str, visit_id: str, visit: VisitCreate, user: dict = Depends(verify_token)):
    result = await db.patients.update_one(
        {"id": patient_id, "visits.id": visit_id},
        {
            "$set": {
                "visits.$.date": visit.date,
                "visits.$.type": visit.type,
                "visits.$.notes": visit.notes,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Patient or visit not found")
    return {"message": "Visit updated"}

@api_router.delete("/patients/{patient_id}/visits/{visit_id}")
async def delete_visit(patient_id: str, visit_id: str, user: dict = Depends(verify_token)):
    result = await db.patients.update_one(
        {"id": patient_id},
        {
            "$pull": {"visits": {"id": visit_id}},
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        }
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Patient not found")
    return {"message": "Visit deleted"}

# ==================== PHOTO ROUTES ====================

@api_router.post("/patients/{patient_id}/visits/{visit_id}/photos", response_model=Photo)
async def add_photo(patient_id: str, visit_id: str, photo: PhotoUpload, user: dict = Depends(verify_token)):
    photo_obj = Photo(**photo.model_dump())
    result = await db.patients.update_one(
        {"id": patient_id, "visits.id": visit_id},
        {
            "$push": {"visits.$.photos": photo_obj.model_dump()},
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        }
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Patient or visit not found")
    return photo_obj

@api_router.delete("/patients/{patient_id}/visits/{visit_id}/photos/{photo_id}")
async def delete_photo(patient_id: str, visit_id: str, photo_id: str, user: dict = Depends(verify_token)):
    result = await db.patients.update_one(
        {"id": patient_id, "visits.id": visit_id},
        {
            "$pull": {"visits.$.photos": {"id": photo_id}},
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        }
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Patient or visit not found")
    return {"message": "Photo deleted"}

# ==================== STATISTICS ROUTES ====================

@api_router.get("/stats", response_model=StatsResponse)
async def get_stats(
    year: Optional[int] = None,
    user: dict = Depends(verify_token)
):
    if not year:
        year = datetime.now().year
    
    # Total patients
    total_patients = await db.patients.count_documents({})
    
    # By status
    status_pipeline = [
        {"$group": {"_id": "$status", "count": {"$sum": 1}}}
    ]
    status_results = await db.patients.aggregate(status_pipeline).to_list(10)
    by_status = {item["_id"]: item["count"] for item in status_results}
    
    # Procedures by month (operated patients with surgery_date in year)
    month_pipeline = [
        {"$match": {"status": "operated", "surgery_date": {"$regex": f"^{year}"}}},
        {"$addFields": {"month": {"$substr": ["$surgery_date", 5, 2]}}},
        {"$group": {"_id": "$month", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}
    ]
    month_results = await db.patients.aggregate(month_pipeline).to_list(12)
    procedures_by_month = [{"month": item["_id"], "count": item["count"]} for item in month_results]
    
    # Procedures by location
    location_pipeline = [
        {"$match": {"status": "operated"}},
        {"$group": {"_id": "$location_id", "count": {"$sum": 1}}}
    ]
    location_results = await db.patients.aggregate(location_pipeline).to_list(100)
    
    # Get location names
    locations = await db.locations.find({}, {"_id": 0}).to_list(100)
    location_map = {loc["id"]: loc["name"] for loc in locations}
    procedures_by_location = [
        {"location": location_map.get(item["_id"], "Unknown"), "count": item["count"]}
        for item in location_results if item["_id"]
    ]
    
    # Revenue by month
    revenue_pipeline = [
        {"$match": {"status": "operated", "surgery_date": {"$regex": f"^{year}"}, "price": {"$exists": True, "$ne": None}}},
        {"$addFields": {"month": {"$substr": ["$surgery_date", 5, 2]}}},
        {"$group": {"_id": "$month", "revenue": {"$sum": "$price"}}},
        {"$sort": {"_id": 1}}
    ]
    revenue_results = await db.patients.aggregate(revenue_pipeline).to_list(12)
    revenue_by_month = [{"month": item["_id"], "revenue": item["revenue"]} for item in revenue_results]
    
    return StatsResponse(
        total_patients=total_patients,
        by_status=by_status,
        procedures_by_month=procedures_by_month,
        procedures_by_location=procedures_by_location,
        revenue_by_month=revenue_by_month
    )

# ==================== EXPORT ROUTES ====================

@api_router.get("/export/patients")
async def export_patients(
    status: Optional[str] = None,
    year: Optional[int] = None,
    user: dict = Depends(verify_token)
):
    query = {}
    if status:
        query["status"] = status
    if year:
        query["surgery_date"] = {"$regex": f"^{year}"}
    
    patients = await db.patients.find(query, {"_id": 0}).to_list(10000)
    locations = await db.locations.find({}, {"_id": 0}).to_list(100)
    location_map = {loc["id"]: loc["name"] for loc in locations}
    
    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"
    
    # Headers
    headers = ["First Name", "Last Name", "Email", "Phone", "Status", "Procedure", "Surgery Date", "Location", "Price", "Notes"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
    
    # Data
    for row, patient in enumerate(patients, 2):
        ws.cell(row=row, column=1, value=patient.get("first_name", ""))
        ws.cell(row=row, column=2, value=patient.get("last_name", ""))
        ws.cell(row=row, column=3, value=patient.get("email", ""))
        ws.cell(row=row, column=4, value=patient.get("phone", ""))
        ws.cell(row=row, column=5, value=patient.get("status", ""))
        ws.cell(row=row, column=6, value=patient.get("procedure_type", ""))
        ws.cell(row=row, column=7, value=patient.get("surgery_date", ""))
        ws.cell(row=row, column=8, value=location_map.get(patient.get("location_id"), ""))
        ws.cell(row=row, column=9, value=patient.get("price", ""))
        ws.cell(row=row, column=10, value=patient.get("notes", ""))
    
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"patients_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== DASHBOARD ROUTES ====================

@api_router.get("/dashboard")
async def get_dashboard(user: dict = Depends(verify_token)):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # All upcoming surgeries (from today onwards)
    upcoming = await db.patients.find(
        {"surgery_date": {"$gte": today}, "status": {"$in": ["planned", "awaiting"]}},
        {"_id": 0}
    ).sort("surgery_date", 1).to_list(100)
    
    # Recent patients
    recent = await db.patients.find({}, {"_id": 0}).sort("created_at", -1).to_list(5)
    
    # Quick stats
    total = await db.patients.count_documents({})
    operated = await db.patients.count_documents({"status": "operated"})
    planned = await db.patients.count_documents({"status": "planned"})
    awaiting = await db.patients.count_documents({"status": "awaiting"})
    consultation = await db.patients.count_documents({"status": "consultation"})
    
    return {
        "upcoming_surgeries": upcoming,
        "recent_patients": recent,
        "stats": {
            "total": total,
            "operated": operated,
            "planned": planned,
            "awaiting": awaiting,
            "consultation": consultation
        }
    }

# ==================== GOOGLE CALENDAR (PLACEHOLDER) ====================

@api_router.get("/calendar/status")
async def calendar_status(user: dict = Depends(verify_token)):
    google_client_id = os.environ.get('GOOGLE_CLIENT_ID')
    return {
        "configured": bool(google_client_id),
        "message": "Google Calendar integration ready" if google_client_id else "Google Calendar not configured. Add GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET to enable."
    }

# ==================== HEALTH CHECK ====================

@api_router.get("/")
async def root():
    return {"message": "AestheticaMD API", "version": "1.0.0"}

# Include the router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
